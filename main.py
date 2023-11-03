from PyQt5 import QtCore, QtGui, QtWidgets

from dotenv import load_dotenv
import time
import json
import datetime
import requests
import pprint
import time
from pathlib import Path
import openpyxl
import sys
import os

import random
import pyperclip


class OCRThread(QtCore.QThread):
    # signal to update QWidgetTable
    update_signal = QtCore.pyqtSignal(object, str)

    def __init__(self, code_bill):
        super().__init__()
        self.code_bill = code_bill

    def run(self):
        cookies = {
            'zalo_id': f'{zaloid}',
            'zlp_token': f'{token}',
            'has_device_id': '0',
        }

        headers = {
            'Host': 'zlp-bill-core-api.zalopay.vn',
            'Accept': 'application/json, text/plain, */*',
            'Origin': 'https://social.zalopay.vn',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 10; MI 8 Lite Build/QQ3A.200805.001; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/74.0.3729.186 Mobile Safari/537.36 ZaloPay Android / 728183',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Referer': f'https://social.zalopay.vn/spa/v2/bill-electric/detail?customercode={self.code_bill}',
            'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
            'X-Requested-With': 'vn.com.vng.zalopay',
        }
        dataf = json.dumps({
            "zlpcustomercode": "",
            "zalopayid": f"{userid}",
            "customercode": self.code_bill,
            "appid": 17}
        )
        now = datetime.datetime.now()
        reqtime = str(int(now.timestamp() * 1000))

        data = {
            'reqdate': reqtime,
            'data': dataf,
        }
        random_time = random.uniform(0.2, 0.7)
        time.sleep(random_time)
        response = requests.post(
            'https://zlp-bill-core-api.zalopay.vn/cpscore/app/getbillinfo',
            cookies=cookies,
            headers=headers,
            data=data,
            verify=False,
        )
        try:
            response.raise_for_status()
        except:
            print(f'Lỗi gọi mã {self.code_bill}')

        result = response.json()

        data = result['data']
        data = json.loads(data)
        # return data
        # pprint.pprint(data)
        self.update_signal.emit(data, self.code_bill)


class Ui_MainWindow(object):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1113, 740)

        # Đường dẫn đến thư mục chứa ứng dụng sau khi đã đóng gói
        app_dir = sys._MEIPASS if getattr(
            sys, 'frozen', False) else os.path.dirname(__file__)

        # Đường dẫn đến tệp icon.ico
        icon_path = os.path.join(app_dir, 'icon.ico')

        with open(icon_path, 'rb') as file:
            icon_data = file.read()

        icon = QtGui.QIcon()
        pixmap = QtGui.QPixmap()
        pixmap.loadFromData(icon_data)
        icon.addPixmap(pixmap)

        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        # font
        font = QtGui.QFont()

        font.setPointSize(12)
        self.edit_text_bill = QtWidgets.QPlainTextEdit(self.centralwidget)
        self.edit_text_bill.setGeometry(QtCore.QRect(20, 60, 181, 381))
        self.edit_text_bill.setObjectName("edit_text_bill")
        self.edit_text_bill.setFont(font)

        font.setPointSize(14)
        self.label_bill = QtWidgets.QLabel(self.centralwidget)
        self.label_bill.setGeometry(QtCore.QRect(20, 20, 191, 31))
        self.label_bill.setFont(font)
        self.label_bill.setObjectName("label_bill")

        font.setPointSize(12)
        self.start_btn = QtWidgets.QPushButton(self.centralwidget)
        self.start_btn.setGeometry(QtCore.QRect(20, 660, 81, 41))
        self.start_btn.setObjectName("start_btn")
        self.start_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.start_btn.setStyleSheet('''
        #start_btn {
            background-color: #7ED86A;  
            border-radius: 3px;
            }
        #start_btn:hover {
            background-color: #42D02D;
        }
        ''')
        self.start_btn.setFont(font)

        self.stop_btn = QtWidgets.QPushButton(self.centralwidget)
        self.stop_btn.setGeometry(QtCore.QRect(120, 660, 85, 41))
        self.stop_btn.setObjectName("stop_btn")
        self.stop_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.stop_btn.setFont(font)
        self.stop_btn.setStyleSheet('''
        #stop_btn {
            background-color: #FF5656;
            border-radius: 3px;
            }
        #stop_btn:hover {
            background-color: #FF0000;
            }
        ''')

        self.reset_btn = QtWidgets.QPushButton(self.centralwidget)
        self.reset_btn.setGeometry(QtCore.QRect(220, 660, 100, 41))
        self.reset_btn.setObjectName("reset_btn")
        self.reset_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.reset_btn.setFont(font)

        self.result_btn = QtWidgets.QPushButton(self.centralwidget)
        self.result_btn.setGeometry(QtCore.QRect(960, 660, 131, 41))
        self.result_btn.setObjectName("result_btn")
        self.result_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.result_btn.setFont(font)

        self.duplicate_btn = QtWidgets.QPushButton(self.centralwidget)
        self.duplicate_btn.setGeometry(QtCore.QRect(20, 460, 161, 41))
        self.duplicate_btn.setObjectName("duplicate_btn")
        self.duplicate_btn.setFont(font)

        self.space_btn = QtWidgets.QPushButton(self.centralwidget)
        self.space_btn.setGeometry(QtCore.QRect(20, 520, 161, 41))
        self.space_btn.setObjectName("space_btn")
        self.space_btn.setFont(font)

        self.debt_btn = QtWidgets.QPushButton(self.centralwidget)
        self.debt_btn.setGeometry(QtCore.QRect(20, 580, 161, 41))
        self.debt_btn.setObjectName("debt_btn")
        self.debt_btn.setFont(font)

        # ------------------- tab ---------------------
        self.title_list = ['STT', 'Mã Hoá Đơn', 'Tổng Nợ', 'Họ Tên', 'Địa Chỉ']
        # them vao list
        result_list.append(self.title_list)
        debt_list.append(self.title_list)
        no_debt_list.append(self.title_list)
        error_list.append(self.title_list[:2])

        font.setPointSize(12)
        # tab_widget
        self.result_tab = QtWidgets.QTabWidget(self.centralwidget)
        self.result_tab.setGeometry(QtCore.QRect(220, 40, 871, 521))
        self.result_tab.setFont(font)
        self.result_tab.setObjectName("result_tab")
        # page_1
        self.total_tab = QtWidgets.QWidget()
        self.total_tab.setObjectName("total_tab")
        # add table_bill vào page 1
        self.table_bill = QtWidgets.QTableWidget(self.total_tab)
        self.table_bill.setGeometry(QtCore.QRect(0, 0, 871, 501))
        self.table_bill.setAlternatingRowColors(False)
        self.table_bill.setRowCount(0)
        self.table_bill.setColumnCount(5)
        self.table_bill.setObjectName("table_bill")
        self.table_bill.setColumnWidth(0, 20)
        self.table_bill.setColumnWidth(1, 150)
        self.table_bill.setColumnWidth(2, 150)
        self.table_bill.setColumnWidth(3, 200)
        self.table_bill.setColumnWidth(4, 370)
        self.table_bill.setHorizontalHeaderLabels(self.title_list)
        self.table_bill.verticalHeader().setVisible(False)
        # page_2
        self.debt_tab = QtWidgets.QWidget()
        self.debt_tab.setObjectName("debt_tab")
        # thêm debt_table vào page 2
        self.debt_table = QtWidgets.QTableWidget(self.debt_tab)
        self.debt_table.setGeometry(QtCore.QRect(0, 0, 871, 501))
        self.debt_table.setAlternatingRowColors(False)
        self.debt_table.setRowCount(0)
        self.debt_table.setColumnCount(5)
        self.debt_table.setObjectName("debt_table")
        self.debt_table.setColumnWidth(0, 20)
        self.debt_table.setColumnWidth(1, 150)
        self.debt_table.setColumnWidth(2, 150)
        self.debt_table.setColumnWidth(3, 200)
        self.debt_table.setColumnWidth(4, 370)
        self.debt_table.setHorizontalHeaderLabels(self.title_list)
        self.debt_table.verticalHeader().setVisible(False)
        # page 3
        self.no_debt_tab = QtWidgets.QWidget()
        self.no_debt_tab.setObjectName("no_debt_tab")
        # thêm no_debt_table vào page 3
        self.no_debt_table = QtWidgets.QTableWidget(self.no_debt_tab)
        self.no_debt_table.setGeometry(QtCore.QRect(0, 0, 871, 501))
        self.no_debt_table.setAlternatingRowColors(False)
        self.no_debt_table.setRowCount(0)
        self.no_debt_table.setColumnCount(5)
        self.no_debt_table.setObjectName("no_debt_table")
        self.no_debt_table.setColumnWidth(0, 20)
        self.no_debt_table.setColumnWidth(1, 150)
        self.no_debt_table.setColumnWidth(2, 150)
        self.no_debt_table.setColumnWidth(3, 200)
        self.no_debt_table.setColumnWidth(4, 370)
        self.no_debt_table.setHorizontalHeaderLabels(self.title_list)
        self.no_debt_table.verticalHeader().setVisible(False)
        # page 4
        self.error_tab = QtWidgets.QWidget()
        self.error_tab.setObjectName("error_tab")
        # thêm error_table vào page 4
        self.error_table = QtWidgets.QTableWidget(self.error_tab)
        self.error_table.setGeometry(QtCore.QRect(0, 0, 871, 501))
        self.error_table.setAlternatingRowColors(False)
        self.error_table.setRowCount(0)
        self.error_table.setColumnCount(2)
        self.error_table.setObjectName("error_table")
        self.error_table.setColumnWidth(0, 20)
        self.error_table.setColumnWidth(1, 150)
        self.error_table.setHorizontalHeaderLabels(self.title_list[:2])
        self.error_table.verticalHeader().setVisible(False)
        # thêm page vào tab
        self.result_tab.addTab(self.total_tab, "Tất Cả")
        self.result_tab.addTab(self.debt_tab, "Nợ Cước")
        self.result_tab.addTab(self.no_debt_tab, "Không Nợ")
        self.result_tab.addTab(self.error_tab, "Mã Lỗi")
        # ------------------- tab ---------------------

        # label result
        font.setPointSize(14)
        self.horizontalLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.horizontalLayoutWidget.setGeometry(
            QtCore.QRect(220, 570, 871, 71))
        self.horizontalLayoutWidget.setObjectName("horizontalLayoutWidget")
        self.horizontalLayoutWidget.setStyleSheet('background-color: #ABC2D0;')
        self.horizontalLayoutWidget.setFont(font)
        self.h_bill_box = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget)
        self.h_bill_box.setContentsMargins(0, 0, 0, 0)
        self.h_bill_box.setObjectName("h_bill_box")

        self.label_total_bill = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_total_bill.setObjectName("label_total_bill")
        self.label_total_bill.setStyleSheet("margin-left: 4px")
        self.label_total_bill.setFont(font)
        self.h_bill_box.addWidget(self.label_total_bill)
        self.h_bill_box.addStretch(20)

        self.label_fail_bill = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_fail_bill.setObjectName("label_fail_bill")
        self.label_fail_bill.setFont(font)
        self.h_bill_box.addWidget(self.label_fail_bill)
        self.h_bill_box.addStretch(15)

        self.label_success_bill = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_success_bill.setObjectName("label_success_bill")
        self.label_success_bill.setFont(font)
        self.h_bill_box.addWidget(self.label_success_bill)
        self.h_bill_box.addStretch(15)

        self.label_error = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_error.setObjectName("label_error")
        self.label_error.setFont(font)
        self.h_bill_box.addWidget(self.label_error)
        self.h_bill_box.addStretch(15)

        self.label_all_bill = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_all_bill.setObjectName("label_all_bill")
        self.label_all_bill.setFont(font)
        self.h_bill_box.addWidget(self.label_all_bill)
        self.h_bill_box.addStretch(35)

        # copy
        self.copy_shortcut = QtWidgets.QShortcut(
            QtGui.QKeySequence.Copy, self.centralwidget)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1131, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        # MainWindow.setWindowTitle(_translate(
        #     "MainWindow", "Tool Check Bill ZaloPay V1.0 By Toan"))
        MainWindow.setWindowTitle(_translate(
            "MainWindow", "TOOL CHECK THẮNG NGỌC 0984402236"))
        self.start_btn.setText(_translate("MainWindow", "Kiểm Tra"))
        self.label_bill.setText(_translate("MainWindow", "Nhập mã hoá đơn :"))
        self.stop_btn.setText(_translate("MainWindow", "Tạm Dừng"))
        self.result_btn.setText(_translate("MainWindow", "Xuất Excel"))
        self.reset_btn.setText(_translate("MainWindow", "Xoá Tất Cả"))
        self.label_total_bill.setText(
            _translate("MainWindow", "Tổng số đơn : "))
        self.label_fail_bill.setText(
            _translate("MainWindow", "Không Nợ Cước :"))
        self.label_success_bill.setText(
            _translate("MainWindow", "Nợ cước :"))
        self.label_error.setText(_translate("MainWindow", "Lỗi :"))
        self.label_all_bill.setText(_translate("MainWindow", "Tổng cước :"))
        self.duplicate_btn.setText(_translate("MainWindow", "Lọc Mã Trùng"))
        self.space_btn.setText(_translate("MainWindow", "Lọc Khoảng Cách"))
        self.debt_btn.setText(_translate("MainWindow", "Lọc Mã Nợ"))

        # start
        self.start_btn.clicked.connect(self.check)
        # stop
        self.stop_btn.clicked.connect(self.stop)
        # reset
        self.reset_btn.clicked.connect(self.reset_input)
        # result
        self.result_btn.clicked.connect(self.export_result)
        # duplicate
        self.duplicate_btn.clicked.connect(self.filter_duplicate)
        # space
        self.space_btn.clicked.connect(self.filter_space)
        # debt
        self.debt_btn.clicked.connect(self.filter_debt)
        # handle_tab
        self.result_tab.tabBarClicked.connect(self.handle_tab)
        # Set Ctrl+C shortcut to trigger copy action
        self.copy_shortcut.activated.connect(self.copy_selected)
    # start

    def check(self):
        # reset index_tab
        self.index_tab = 0
        # reset table
        self.table_bill.clearContents()
        self.table_bill.setRowCount(0)
        self.debt_table.clearContents()
        self.debt_table.setRowCount(0)
        self.no_debt_table.clearContents()
        self.no_debt_table.setRowCount(0)
        self.error_table.clearContents()
        self.error_table.setRowCount(0)
        # reset list
        result_list.clear()
        debt_list.clear()
        no_debt_list.clear()
        error_list.clear()
        result_list.append(self.title_list)
        debt_list.append(self.title_list)
        no_debt_list.append(self.title_list)
        error_list.append(self.title_list[:2])
        # reset label
        self.label_total_bill.setText('Tổng số đơn :')
        self.label_fail_bill.setText('Không Nợ Cước :')
        self.label_success_bill.setText('Nợ Cước :')
        self.label_error.setText('Lỗi :')
        self.label_all_bill.setText('Tổng cước :')

        self.start_btn.setEnabled(False)
        self.count_thread = 0
        self.temp = True
        self.all_bill_count = 0
        self.fail_count = 0
        self.error_count = 0
        self.success_count = 0
        self.debt_code_list = []
        self.edit_text_list = self.edit_text_bill.toPlainText().split('\n')
        self.thread_pool = []
        for pos, edit_text_item in enumerate(self.edit_text_list):
            if not self.temp:
                break
            if not edit_text_item.strip():
                QtWidgets.QMessageBox.warning(
                    self.centralwidget, 'Thông báo', 'Lỗi ! Mời lọc khoảng cách và mã trùng để tiếp tục !!!')
                self.reset_input()
                break
            thread = OCRThread(edit_text_item)
            thread.update_signal.connect(self.update_result)
            self.thread_pool.append(thread)
            thread.start()
            thread.finished.connect(lambda: self.add_count_thread())

    def add_count_thread(self):
        self.count_thread += 1
        if self.count_thread == len(self.edit_text_list):
            # start
            self.start_btn.setEnabled(True)
            # table
            self.table_bill.verticalHeader().setStyleSheet(
                'QHeaderView::section {background-color: #FFE5B8; }')
            self.table_bill.setVerticalHeaderLabels([' '] * len(result_list))
            self.table_bill.verticalHeader().setVisible(True)

            self.debt_table.verticalHeader().setStyleSheet(
                'QHeaderView::section {background-color: #FFE5B8; }')
            self.debt_table.setVerticalHeaderLabels([' '] * len(debt_list))
            self.debt_table.verticalHeader().setVisible(True)

            self.no_debt_table.verticalHeader().setStyleSheet(
                'QHeaderView::section {background-color: #FFE5B8; }')
            self.no_debt_table.setVerticalHeaderLabels(
                [' '] * len(no_debt_list))
            self.no_debt_table.verticalHeader().setVisible(True)

            self.error_table.verticalHeader().setStyleSheet(
                'QHeaderView::section {background-color: #FFE5B8; }')
            self.error_table.setVerticalHeaderLabels([' '] * len(error_list))
            self.error_table.verticalHeader().setVisible(True)
            # show result
            self.show_result()

    # use for start
    def update_result(self, data, code_bill):
        if data['appid'] == 0:
            # QtWidgets.QMessageBox.warning(
            #     self.centralwidget, 'Thông báo', f'Lỗi ! Mã đơn {code_bill} không hợp lệ !!!')
            error_row = self.error_table.rowCount()
            error_list.append([error_row+1, code_bill])
            self.error_table.insertRow(error_row)
            index_error = QtWidgets.QTableWidgetItem(str(error_row + 1))
            code_error = QtWidgets.QTableWidgetItem(code_bill)
            self.error_table.setItem(error_row, 0, index_error)
            self.error_table.setItem(error_row, 1, code_error)
            self.error_count += 1
            return
        # quet du lieu
        # lay du lieu
        list_bill = data['billlist']
        if not list_bill:
            debt_all = 0
            debt_all_text = '0'
            self.fail_count += 1
        else:
            self.debt_code_list.append(code_bill)
            debt_all = data['totalamount']
            # tinh tong
            self.all_bill_count += debt_all
            debt_all_text = '{:,d}'.format(debt_all)

        name_text = '' if not 'customername' in data else data['customername']
        address_text = '' if not 'address' in data else data['address']

        table_row = self.table_bill.rowCount()
        debt_row = self.debt_table.rowCount()
        no_debt_row = self.no_debt_table.rowCount()
        # them vao debt_list
        if list_bill:
            debt_list.append([debt_row + 1, code_bill,
                              debt_all, name_text, address_text])
        else:
            no_debt_list.append([no_debt_row + 1, code_bill,
                                 debt_all, name_text, address_text])
        # them vao result_list
        result_list.append([table_row + 1, code_bill,
                            debt_all, name_text, address_text])

        # thêm vào table_bill
        self.table_bill.insertRow(table_row)
        self.add_table(table_row, code_bill, debt_all_text,
                       name_text, address_text)
        self.table_bill.setItem(table_row, 0, self.index_item)
        self.table_bill.setItem(table_row, 1, self.code_bill_item)
        self.table_bill.setItem(table_row, 2, self.debt_all_item)
        self.table_bill.setItem(table_row, 3, self.name_text_item)
        self.table_bill.setItem(table_row, 4, self.address_text_item)
        # thêm vào debt_table
        if list_bill:
            self.debt_table.insertRow(debt_row)
            self.add_table(debt_row, code_bill, debt_all_text,
                           name_text, address_text)
            self.debt_table.setItem(debt_row, 0, self.index_item)
            self.debt_table.setItem(debt_row, 1, self.code_bill_item)
            self.debt_table.setItem(debt_row, 2, self.debt_all_item)
            self.debt_table.setItem(debt_row, 3, self.name_text_item)
            self.debt_table.setItem(debt_row, 4, self.address_text_item)
        else:
            self.no_debt_table.insertRow(no_debt_row)
            self.add_table(no_debt_row, code_bill,
                           debt_all_text, name_text, address_text)
            self.no_debt_table.setItem(no_debt_row, 0, self.index_item)
            self.no_debt_table.setItem(no_debt_row, 1, self.code_bill_item)
            self.no_debt_table.setItem(no_debt_row, 2, self.debt_all_item)
            self.no_debt_table.setItem(no_debt_row, 3, self.name_text_item)
            self.no_debt_table.setItem(no_debt_row, 4, self.address_text_item)

    def add_table(self, row, code_bill, debt_all_text, name_text, address_text):
        # set item
        self.index_item = QtWidgets.QTableWidgetItem(str(row + 1))
        self.code_bill_item = QtWidgets.QTableWidgetItem(code_bill)
        self.debt_all_item = QtWidgets.QTableWidgetItem(debt_all_text)
        self.name_text_item = QtWidgets.QTableWidgetItem(name_text)
        self.address_text_item = QtWidgets.QTableWidgetItem(address_text)
        # can giua
        self.index_item.setTextAlignment(0x0004 | 0x0080)
        # self.code_bill_item.setTextAlignment(0x0004 | 0x0080)
        self.index_item.setTextAlignment(0x0004 | 0x0080)
        self.debt_all_item.setTextAlignment(0x0004 | 0x0080)
        # self.name_text_item.setTextAlignment(0x0004 | 0x0080)
        # self.address_text_item.setTextAlignment(0x0004 | 0x0080)

    # reset
    def reset_input(self):
        # reset input
        self.edit_text_bill.setPlainText('')
        # reset table
        self.table_bill.clearContents()
        self.table_bill.setRowCount(0)
        self.debt_table.clearContents()
        self.debt_table.setRowCount(0)
        self.no_debt_table.clearContents()
        self.no_debt_table.setRowCount(0)
        self.error_table.clearContents()
        self.error_table.setRowCount(0)

        # reset list
        result_list.clear()
        debt_list.clear()
        no_debt_list.clear()
        error_list.clear()
        result_list.append(self.title_list)
        debt_list.append(self.title_list)
        no_debt_list.append(self.title_list)
        error_list.append(self.title_list[:2])
        # reset label
        self.label_total_bill.setText('Tổng số đơn :')
        self.label_fail_bill.setText('Không Nợ Cước :')
        self.label_success_bill.setText('Nợ Cước :')
        self.label_error.setText('Lỗi :')
        self.label_all_bill.setText('Tổng cước :')
        self.start_btn.setEnabled(True)

    # stop

    def stop(self):
        self.temp = False
        for thread in self.thread_pool:
            if thread.isRunning():
                thread.terminate()
        self.start_btn.setEnabled(True)
    # handle Tab

    def handle_tab(self, index):
        self.index_tab = index

    # copy
    def copy_selected(self):
        try:
            if self.index_tab == 0:
                self.table = self.table_bill
            elif self.index_tab == 1:
                self.table = self.debt_table
            elif self.index_tab == 2:
                self.table = self.no_debt_table
            elif self.index_tab == 3:
                self.table = self.error_table
        except:
            print('Lỗi copy')
            return
        selected_indexes = self.table.selectedIndexes()
        if selected_indexes:
            rows = list(set(index.row() for index in selected_indexes))
            cols = list(set(index.column() for index in selected_indexes))

            data = ""
            for row in rows:
                for col in cols:
                    item = self.table.item(row, col)
                    data += f"{item.text()}\t"  # Use tabs to separate columns

                data = data[:-1]  # Remove the last tab for the row
                data += "\n"  # Move to the next row

            pyperclip.copy(data)
            print("Selected cells copied to clipboard.")
    # export

    def export_result(self):
        try:
            if self.index_tab:
                pass
        except:
            QtWidgets.QMessageBox.warning(
                self.centralwidget, 'Thông báo', 'Vui lòng nhập hoá đơn để check !!!')
            return

        time_now = datetime.datetime.now()
        time_convert = time_now.strftime('%d-%m-%Y %H\'%M\'%S')
        wb = openpyxl.Workbook()
        sheet = wb.active
        # chinh do dai
        sheet.column_dimensions['B'].width = 17
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 56

        if self.index_tab == 0:
            file_path = desktop_path / f'Hoá Đơn {time_convert}.xlsx'
            # ghi file
            for index, result_item in enumerate(result_list):
                sheet.cell(row=index + 1, column=1).value = result_item[0]
                sheet.cell(row=index + 1, column=2).value = result_item[1]
                sheet.cell(row=index + 1, column=3).value = result_item[2]
                sheet.cell(row=index + 1, column=3).number_format = '#,##0'
                sheet.cell(row=index + 1, column=4).value = result_item[3]
                sheet.cell(row=index + 1, column=5).value = result_item[4]
            # luu file
            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self.centralwidget, 'Thông báo', F'Xuất hoá đơn {time_convert} thành công !!!')
        elif self.index_tab == 1:
            file_path = desktop_path / f'Nợ cước {time_convert}.xlsx'
            # ghi file
            for index, debt_item in enumerate(debt_list):
                sheet.cell(row=index + 1, column=1).value = debt_item[0]
                sheet.cell(row=index + 1, column=2).value = debt_item[1]
                sheet.cell(row=index + 1, column=3).value = debt_item[2]
                sheet.cell(row=index + 1, column=3).number_format = '#,##0'
                sheet.cell(row=index + 1, column=4).value = debt_item[3]
                sheet.cell(row=index + 1, column=5).value = debt_item[4]
            # luu file
            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self.centralwidget, 'Thông báo', F'Xuất Nợ Cước {time_convert} thành công !!!')
        elif self.index_tab == 2:
            file_path = desktop_path / f'Không Nợ {time_convert}.xlsx'
            # ghi file
            for index, no_debt_item in enumerate(no_debt_list):
                sheet.cell(row=index + 1, column=1).value = no_debt_item[0]
                sheet.cell(row=index + 1, column=2).value = no_debt_item[1]
                sheet.cell(row=index + 1, column=3).value = no_debt_item[2]
                sheet.cell(row=index + 1, column=3).number_format = '#,##0'
                sheet.cell(row=index + 1, column=4).value = no_debt_item[3]
                sheet.cell(row=index + 1, column=5).value = no_debt_item[4]
            # luu file
            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self.centralwidget, 'Thông báo', F'Xuất Không Nợ {time_convert} thành công !!!')
        elif self.index_tab == 3:
            file_path = desktop_path / f'Mã Lỗi {time_convert}.xlsx'
            # ghi file
            for index, error_item in enumerate(error_list):
                sheet.cell(row=index + 1, column=1).value = error_item[0]
                sheet.cell(row=index + 1, column=2).value = error_item[1]
            # luu file
            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self.centralwidget, 'Thông báo', F'Xuất Mã Lỗi {time_convert} thành công !!!')
    # filter duplicate

    def filter_duplicate(self):
        self.edit_text_list = list(dict.fromkeys(
            self.edit_text_bill.toPlainText().split('\n')))
        self.edit_text_bill.setPlainText('\n'.join(self.edit_text_list))

    # filter space

    def filter_space(self):
        # self.edit_text_list = [item.strip().split(
        #     '\t')[0] for item in self.edit_text_bill.toPlainText().split('\n')]
        # self.edit_text_list = [item.strip().split(
        #     ' ')[0] for item in self.edit_text_list]
        self.edit_text_list = [item.strip(
        )[:13] for item in self.edit_text_bill.toPlainText().split('\n') if item.strip()]
        self.edit_text_bill.setPlainText('\n'.join(self.edit_text_list))

    # filter debt

    def filter_debt(self):
        self.edit_text_bill.setPlainText('\n'.join(self.debt_code_list))

    # show data result
    def show_result(self):
        self.all_bill_count = '{:,d}'.format(self.all_bill_count)
        self.label_total_bill.setText(
            'Tổng số đơn : %s' % (len(result_list) - 1 + self.error_count))
        self.label_fail_bill.setText('Không Nợ Cước : %s' % self.fail_count)
        self.label_success_bill.setText('Nợ Cước : %s' % (
            len(result_list) - 1 - self.fail_count))
        self.label_error.setText('Lỗi : %s' % self.error_count)
        self.label_all_bill.setText('Tổng cước : %s' % self.all_bill_count)


if __name__ == "__main__":
    import sys
    load_dotenv()
    zaloid = os.getenv('ZALOID')
    userid = os.getenv('USERID')
    token = os.getenv('TOKEN')
    result_list = []
    debt_list = []
    no_debt_list = []
    error_list = []
    desktop_path = Path.home() / "Desktop"
    # time_next_main = datetime.datetime(2023, 11, 1, 18, 0, 0)
    # time_now_main = datetime.datetime.now()
    # if time_next_main > time_now_main:
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
    # else:
    #     print('Hết thời hạn thử')
